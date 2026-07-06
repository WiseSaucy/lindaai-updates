#!/usr/bin/env python3
"""
LindaAI — RV Park auto-fill.

Takes a JSON of deal data (extracted from a seller's T-12 / P&L / OM) and writes
it into the RV_Park_Underwriting.xlsx template:

  • Property basics, income and financing  -> Underwriting tab
  • Seller's REPORTED expenses              -> Normalization tab (seller column)
  • NORMALIZED expenses (Linda's 4 rules)       -> Underwriting tab (so NOI, the
    Deal Scorecard, and Offer Structures all reflect YOUR numbers, not theirs)

Linda's normalization rules (overridable in the JSON's "normalization" block):
  management -> >= 10% of EGI · repairs -> >= 5% of EGI · capex -> >= 3% of EGI
  · property taxes -> +20%

Usage:
  python3 fill_template.py deal.json --template RV_Park_Underwriting.xlsx --out "Deal - Example RV Park.xlsx"

The JSON schema is documented in deal_input.example.json (same folder).
Everything downstream of the input cells is formula-driven, so it recalculates
the moment the file is opened in Excel / Google Sheets.
"""
import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: openpyxl  (pip install openpyxl)")


def pmt(rate, nper, pv):
    """Excel-equivalent PMT (returns positive payment for negative pv)."""
    if nper == 0:
        return 0.0
    if rate == 0:
        return -pv / nper
    return -(pv * rate) / (1 - (1 + rate) ** (-nper))

# --- cell maps (MUST match build_rv_underwriting.py) ------------------------
U = {  # Underwriting tab
    "name": "B3", "loc": "B4", "date": "B5",
    "sites": "B8", "price": "B9", "rent": "B11", "vac": "B12", "other_m": "B13",
    "mgmt_pct": "B14", "res_site": "B15", "rent_g": "B16", "exp_g": "B17",
    "tax": "B27", "ins": "B28", "util": "B29", "rm": "B30", "pay": "B31", "admin": "B32",
    "ltv": "B44", "intr": "B45", "amort": "B46", "close_pct": "B47", "capex": "B48",
    "hold": "B49", "exitcap": "B50", "sellcost": "B51",
}
# Normalization tab — seller-reported column (B)
N = {"egi": "B11", "tax": "B15", "ins": "B16", "util": "B17", "rm": "B18",
     "mgmt": "B19", "capex": "B20", "pay": "B21", "admin": "B22"}
# Normalization tab — Linda's rule cells (kept in sync with any JSON overrides)
N_RULES = {"mgmt_pct": "B5", "rm_pct": "B6", "capex_pct": "B7", "tax_bump": "B8"}
# expected labels in column A (sanity check that the template hasn't shifted)
U_CHECK = {"B8": "Number of RV Sites", "B9": "Purchase Price", "B27": "Property Taxes"}
N_CHECK = {"B11": "Effective Gross Income", "B15": "Property Taxes"}


def warn(msg):
    print(f"  ! {msg}", file=sys.stderr)


def check_labels(ws, checks, tab):
    for cell, expect in checks.items():
        row = cell[1:]
        got = str(ws[f"A{row}"].value or "")
        if expect.lower() not in got.lower():
            warn(f"{tab}!A{row} = '{got}' (expected to contain '{expect}'). "
                 "Template layout may have changed — verify the filled cells.")


def num(d, *keys, default=0.0):
    cur = d
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur if cur is not None else default


def main():
    ap = argparse.ArgumentParser(description="Fill the RV park underwriting template from deal JSON.")
    ap.add_argument("deal_json")
    ap.add_argument("--template", default=None,
                    help="path to RV_Park_Underwriting.xlsx (default: the copy shipped with this skill)")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    if args.template is None:
        here = os.path.dirname(os.path.abspath(__file__))
        for cand in (os.path.join(here, "RV_Park_Underwriting.xlsx"),
                     "RV_Park_Underwriting.xlsx",
                     os.path.join(here, "..", "..", "RV_Park_Underwriting.xlsx")):
            if os.path.exists(cand):
                args.template = cand
                break
        else:
            sys.exit("Template RV_Park_Underwriting.xlsx not found — pass --template /path/to/it")

    with open(args.deal_json, "r", encoding="utf-8") as fh:
        d = json.load(fh)

    wb = openpyxl.load_workbook(args.template)
    for tab in ("Underwriting", "Normalization"):
        if tab not in wb.sheetnames:
            sys.exit(f"Template is missing the '{tab}' tab — is this the right file?")
    uw, nm = wb["Underwriting"], wb["Normalization"]
    check_labels(uw, U_CHECK, "Underwriting")
    check_labels(nm, N_CHECK, "Normalization")

    # ---- normalization rules (Linda defaults) ----
    mgmt_pct = num(d, "normalization", "mgmt_pct", default=0.10)
    rm_pct   = num(d, "normalization", "rm_pct",   default=0.05)
    capex_pct= num(d, "normalization", "capex_pct",default=0.03)
    tax_bump = num(d, "normalization", "tax_bump", default=0.20)

    # ---- income ----
    sites   = num(d, "sites", default=0) or 1
    price   = num(d, "asking_price")
    gpr     = num(d, "income", "annual_site_rent")       # gross potential site rent (annual)
    vac     = num(d, "income", "vacancy_pct")
    other_a = num(d, "income", "other_income_annual")
    egi     = gpr * (1 - vac) + other_a                  # Linda: gross income = EGI

    # ---- seller-reported expenses ----
    s = d.get("seller_expenses", {})
    s_tax  = num(s, "property_taxes"); s_ins = num(s, "insurance"); s_util = num(s, "utilities")
    s_rm   = num(s, "repairs_maintenance"); s_mgmt = num(s, "management"); s_capex = num(s, "capex")
    s_pay  = num(s, "payroll"); s_admin = num(s, "other_admin")

    # ---- normalized expenses (Linda's rules) ----
    n_tax  = s_tax * (1 + tax_bump)
    n_rm   = max(s_rm,   rm_pct * egi)
    n_capex= max(s_capex,capex_pct * egi)
    eff_mgmt_pct = max(mgmt_pct, (s_mgmt / egi) if egi else mgmt_pct)
    eff_res_site = n_capex / sites                       # model holds reserves per-site

    # ---- write Underwriting (basics + financing + NORMALIZED expenses) ----
    uw[U["name"]] = d.get("property_name", "")
    uw[U["loc"]]  = d.get("location", "")
    if d.get("analysis_date"): uw[U["date"]] = d["analysis_date"]
    uw[U["sites"]] = sites
    uw[U["price"]] = price
    uw[U["rent"]]  = (gpr / sites / 12) if sites else 0   # back into avg monthly rent
    uw[U["vac"]]   = vac
    uw[U["other_m"]] = other_a / 12
    uw[U["mgmt_pct"]] = eff_mgmt_pct
    uw[U["res_site"]] = eff_res_site
    uw[U["tax"]] = n_tax
    uw[U["ins"]] = s_ins
    uw[U["util"]] = s_util
    uw[U["rm"]] = n_rm
    uw[U["pay"]] = s_pay
    uw[U["admin"]] = s_admin
    f = d.get("financing", {})
    for jkey, ukey in (("ltv","ltv"),("interest_rate","intr"),("amortization_years","amort"),
                       ("closing_pct","close_pct"),("initial_capex","capex"),("hold_years","hold"),
                       ("exit_cap","exitcap"),("selling_cost_pct","sellcost")):
        if jkey in f and f[jkey] is not None:
            uw[U[ukey]] = f[jkey]
    if "rent_growth" in d: uw[U["rent_g"]] = d["rent_growth"]
    if "expense_growth" in d: uw[U["exp_g"]] = d["expense_growth"]

    # ---- write Normalization (seller-reported column + rule overrides) ----
    nm[N["egi"]] = egi
    nm[N["tax"]] = s_tax; nm[N["ins"]] = s_ins; nm[N["util"]] = s_util; nm[N["rm"]] = s_rm
    nm[N["mgmt"]] = s_mgmt; nm[N["capex"]] = s_capex; nm[N["pay"]] = s_pay; nm[N["admin"]] = s_admin
    # keep the tab's rule cells in sync with any JSON overrides so the
    # Normalization tab, Underwriting tab and reports all tell the same story
    nm[N_RULES["mgmt_pct"]] = mgmt_pct
    nm[N_RULES["rm_pct"]] = rm_pct
    nm[N_RULES["capex_pct"]] = capex_pct
    nm[N_RULES["tax_bump"]] = tax_bump

    # ---- normalized NOI + financing (used for the MAO seed below) ----
    opex_norm_ = (n_tax + s_ins + s_util + n_rm + (eff_mgmt_pct * egi)
                  + n_capex + s_pay + s_admin)
    noi_norm = egi - opex_norm_
    ltv = f.get("ltv", 0.70) or 0.70
    rate = f.get("interest_rate", 0.07) or 0.07
    amort = f.get("amortization_years", 25) or 25
    close_pct = f.get("closing_pct", 0.03) or 0.03
    icapex = f.get("initial_capex", 0) or 0

    # ---- Offer Structures: seed prices + financing where present ----
    if "Offer Structures" in wb.sheetnames:
        of = wb["Offer Structures"]
        for col in ("B", "D", "E"):    # B=asking, D=partial carry, E=full carry hold asking price
            of[f"{col}18"] = price
        # Offer 1 (Conventional) = the deal's own MAO, not the template's demo number
        k = pmt(rate / 12, amort * 12, -1) * 12
        p_dscr = noi_norm / (1.35 * ltv * k) if (ltv and k) else 0
        denom = ltv * k + 0.10 * (1 - ltv + close_pct)
        p_coc = (noi_norm - 0.10 * icapex) / denom if denom else 0
        mao = max(0, min(p_dscr, p_coc))
        # floor to the nearest $1k — rounding UP would push the offer past its own MAO
        of["C18"] = int(mao // 1000) * 1000 if mao else price
        # keep the MAO block + offer rows on the deal's actual financing
        if f.get("ltv") is not None:
            of["B7"] = f["ltv"]
            of["C19"] = f["ltv"]          # conventional offer bank %
        if f.get("interest_rate") is not None:
            of["B8"] = f["interest_rate"]
            for col in "BCDE": of[f"{col}20"] = f["interest_rate"]
        if f.get("amortization_years") is not None:
            of["B9"] = f["amortization_years"]
            for col in "BCDE": of[f"{col}21"] = f["amortization_years"]

    out = args.out or f"Deal - {d.get('property_name','RV Park')}.xlsx"
    wb.save(out)

    # quick echo so the agent/user can sanity-check
    opex_seller = s_tax + s_ins + s_util + s_rm + s_mgmt + s_capex + s_pay + s_admin
    opex_norm = n_tax + s_ins + s_util + n_rm + (eff_mgmt_pct * egi) + n_capex + s_pay + s_admin
    print(json.dumps({
        "ok": True, "out": out, "sites": sites, "asking_price": price,
        "EGI": round(egi), "seller_NOI": round(egi - opex_seller),
        "normalized_NOI": round(egi - opex_norm),
        "seller_expense_ratio": round(opex_seller / egi, 4) if egi else None,
        "normalized_expense_ratio": round(opex_norm / egi, 4) if egi else None,
        "noi_haircut": round((egi - opex_seller) - (egi - opex_norm)),
    }, indent=2))


if __name__ == "__main__":
    main()
