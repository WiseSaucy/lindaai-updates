#!/usr/bin/env python3
"""Builds an RV Park underwriting workbook (live Excel formulas).

Tabs: Read Me | Deal Summary | Underwriting | Income Detail |
      Actuals vs Pro Forma | Pro Forma & Returns | Sensitivity
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ---- styles ----------------------------------------------------------------
TITLE   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
SECTION = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD    = Font(name="Calibri", size=11, bold=True)
LABEL   = Font(name="Calibri", size=11)
INPUTF  = Font(name="Calibri", size=11, bold=True, color="1F4E78")
NOTE    = Font(name="Calibri", size=9, italic=True, color="808080")
BIG      = Font(name="Calibri", size=20, bold=True, color="1F4E78")

TITLE_FILL   = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
INPUT_FILL   = PatternFill("solid", fgColor="FFF2CC")   # yellow = type here
TOTAL_FILL   = PatternFill("solid", fgColor="D9E1F2")
KEY_FILL     = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL    = PatternFill("solid", fgColor="FCE4D6")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CUR  = '#,##0'
CUR2 = '$#,##0'
PCT  = '0.0%'
PCT2 = '0.00%'
MULT = '0.00"x"'

wb = Workbook()
wb.remove(wb.active)  # drop default empty sheet
R = {}  # Underwriting cell registry: name -> "B17"

# generic helpers -------------------------------------------------------------
def merge_title(ws, rng, text):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[int(rng.split(":")[0][1:])].height = 26

def sec(ws, row, text, span="A:C"):
    a, b = span.split(":")
    ws.merge_cells(f"{a}{row}:{b}{row}")
    c = ws[f"{a}{row}"]; c.value = text; c.font = SECTION; c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20

# =============================================================================
# SHEET — INCOME DETAIL (built first; Underwriting GPR can reference it)
# =============================================================================
inc = wb.create_sheet("Income Detail")
inc.sheet_view.showGridLines = False
inc.column_dimensions["A"].width = 30
for col in "BCDE":
    inc.column_dimensions[col].width = 16
merge_title(inc, "A1:E1", "INCOME DETAIL — SITE MIX & SEASONALITY")
inc["A2"] = "Optional. Build income from your actual site mix, then set 'Use Detailed Income Build?' = Y on the Underwriting tab."
inc["A2"].font = NOTE

def inp(ws, cell, val, fmt, note_cell=None, note=None):
    c = ws[cell]; c.value = val; c.font = INPUTF; c.fill = INPUT_FILL
    c.number_format = fmt; c.border = BORDER; c.alignment = Alignment(horizontal="right")
    if note_cell and note:
        ws[note_cell] = note; ws[note_cell].font = NOTE

def calc(ws, cell, formula, fmt, bold=False, fill=None):
    c = ws[cell]; c.value = formula; c.number_format = fmt
    c.font = BOLD if bold else LABEL; c.alignment = Alignment(horizontal="right")
    if fill: c.fill = fill

# Long-term tenants ----------------------------------------------------------
sec(inc, 4, "LONG-TERM TENANTS (MONTHLY / ANNUAL LEASES)", "A:E")
inc["A5"] = "Tier"; inc["B5"] = "# Sites"; inc["C5"] = "Monthly Rate"
inc["D5"] = "Occupancy %"; inc["E5"] = "Annual Revenue"
for cl in "ABCDE": inc[f"{cl}5"].font = BOLD
inc["A6"] = "Monthly tenants"; inc["A7"] = "Annual / seasonal leases"
for r in (6, 7):
    inp(inc, f"B{r}", 30 if r == 6 else 8, CUR)
    inp(inc, f"C{r}", 450 if r == 6 else 400, CUR2)
    inp(inc, f"D{r}", 0.95 if r == 6 else 0.90, PCT)
    calc(inc, f"E{r}", f"=B{r}*C{r}*12*D{r}", CUR2)
inc["A8"] = "Subtotal — long-term"; inc["A8"].font = BOLD
calc(inc, "E8", "=SUM(E6:E7)", CUR2, bold=True, fill=TOTAL_FILL)
calc(inc, "B8", "=SUM(B6:B7)", CUR, bold=True)

# Transient (nightly) seasonal ----------------------------------------------
sec(inc, 10, "TRANSIENT / NIGHTLY SITES (SEASONAL)", "A:E")
inc["A11"] = "# Transient sites"; inp(inc, "B11", 12, CUR)
inc["A12"] = "Avg nightly rate"; inp(inc, "B12", 55, CUR2)
inc["A14"] = "Month"; inc["B14"] = "Days"; inc["C14"] = "Occupancy %"
inc["D14"] = "Revenue"
for cl in "ABCD": inc[f"{cl}14"].font = BOLD
months = [("Jan",31,.30),("Feb",28,.30),("Mar",31,.45),("Apr",30,.55),
          ("May",31,.70),("Jun",30,.85),("Jul",31,.95),("Aug",31,.95),
          ("Sep",30,.75),("Oct",31,.55),("Nov",30,.35),("Dec",31,.30)]
start = 15
for i,(m,days,occ) in enumerate(months):
    r = start + i
    inc[f"A{r}"] = m; inc[f"A{r}"].font = LABEL
    inc[f"B{r}"] = days; inc[f"B{r}"].font = LABEL; inc[f"B{r}"].alignment = Alignment(horizontal="right")
    inp(inc, f"C{r}", occ, PCT)
    calc(inc, f"D{r}", f"=$B$11*$B$12*B{r}*C{r}", CUR2)
last = start + 11           # 26
trow = last + 1             # 27 transient subtotal
inc[f"A{trow}"] = "Subtotal — transient"; inc[f"A{trow}"].font = BOLD
calc(inc, f"D{trow}", f"=SUM(D{start}:D{last})", CUR2, bold=True, fill=TOTAL_FILL)
arow = trow + 1
inc[f"A{arow}"] = "Avg transient occupancy"; inc[f"A{arow}"].font = LABEL
calc(inc, f"C{arow}", f"=SUMPRODUCT(B{start}:B{last},C{start}:C{last})/SUM(B{start}:B{last})", PCT)

# Total potential annual site revenue ---------------------------------------
tot = arow + 2              # 30
sec(inc, tot - 1, "TOTAL POTENTIAL ANNUAL SITE REVENUE", "A:E")
inc[f"A{tot}"] = "Total Potential Annual Site Revenue"; inc[f"A{tot}"].font = BOLD
calc(inc, f"E{tot}", f"=E8+D{trow}", CUR2, bold=True, fill=KEY_FILL)
INC_TOTAL = f"'Income Detail'!E{tot}"     # referenced by Underwriting GPR
inc[f"A{tot+1}"] = "Total sites in build"
calc(inc, f"E{tot+1}", f"=B8+B11", CUR)
inc[f"A{tot+2}"] = "Implied blended monthly rent / site"
calc(inc, f"E{tot+2}", f"=E{tot}/E{tot+1}/12", CUR2)
inc[f"A{tot+3}"] = ("Note: when Use Detailed Income Build = Y, keep the Underwriting "
                    "Vacancy % low (credit loss only) — occupancy is already in this build.")
inc[f"A{tot+3}"].font = NOTE

# =============================================================================
# SHEET — UNDERWRITING (row cursor; registry R for cross-sheet refs)
# =============================================================================
ws = wb.create_sheet("Underwriting")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 32

merge_title(ws, "A1:C1", "RV PARK UNDERWRITING ANALYSIS")

def u_input(row, label, value, fmt=CUR2, note="", key=None):
    ws[f"A{row}"] = label; ws[f"A{row}"].font = LABEL
    c = ws[f"B{row}"]; c.value = value; c.font = INPUTF; c.fill = INPUT_FILL
    c.number_format = fmt; c.border = BORDER; c.alignment = Alignment(horizontal="right")
    if note: ws[f"C{row}"] = note; ws[f"C{row}"].font = NOTE
    if key: R[key] = f"B{row}"

def u_calc(row, label, formula, fmt=CUR2, note="", bold=False, fill=None, key=None):
    ws[f"A{row}"] = label; ws[f"A{row}"].font = BOLD if bold else LABEL
    c = ws[f"B{row}"]; c.value = formula; c.font = BOLD if bold else LABEL
    c.number_format = fmt; c.alignment = Alignment(horizontal="right")
    if fill: c.fill = fill
    if note: ws[f"C{row}"] = note; ws[f"C{row}"].font = NOTE
    if key: R[key] = f"B{row}"

# header inputs
ws["A3"]="Property Name"; ws["A3"].font=LABEL
ws["B3"]="Enter property name"; ws["B3"].font=INPUTF; ws["B3"].fill=INPUT_FILL; ws["B3"].border=BORDER; R["name"]="B3"
ws["A4"]="Location (City, ST)"; ws["A4"].font=LABEL
ws["B4"]="City, ST"; ws["B4"].font=INPUTF; ws["B4"].fill=INPUT_FILL; ws["B4"].border=BORDER; R["loc"]="B4"
ws["A5"]="Analysis Date"; ws["A5"].font=LABEL
ws["B5"]="2026-06-26"; ws["B5"].font=INPUTF; ws["B5"].fill=INPUT_FILL; ws["B5"].border=BORDER; R["date"]="B5"
ws["C3"]="Yellow cells = your inputs. Everything else calculates."; ws["C3"].font=NOTE

sec(ws, 7, "DEAL ASSUMPTIONS")
u_input(8,  "Number of RV Sites (pads)", 50, CUR, "total rentable pads", key="sites")
u_input(9,  "Purchase Price", 1500000, CUR2, key="price")
u_calc (10, "  Price per Site", "=B9/B8", CUR2, "purchase price / sites")
u_input(11, "Avg. Monthly Rent per Site", 450, CUR2, "blended lot rent / month", key="rent")
u_input(12, "Vacancy & Credit Loss %", 0.15, PCT, "economic vacancy", key="vac")
u_input(13, "Other Income (monthly)", 1500, CUR2, "store, laundry, propane, fees", key="other_m")
u_input(14, "Management Fee (% of EGI)", 0.08, PCT, "3rd-party or self-mgmt", key="mgmt_pct")
u_input(15, "Replacement Reserves / site / yr", 100, CUR2, "capital reserve", key="res_site")
u_input(16, "Annual Rent Growth %", 0.03, PCT, "for multi-year projection", key="rent_g")
u_input(17, "Annual Expense Growth %", 0.025, PCT, "for multi-year projection", key="exp_g")
u_input(18, "Use Detailed Income Build? (Y/N)", "N", '@',
        "Y = pull GPR from Income Detail tab", key="use_detail")

sec(ws, 20, "INCOME  (Year 1 Stabilized Pro Forma)")
u_calc(21, "Gross Potential Rent (GPR)",
       f'=IF({R["use_detail"]}="Y",{INC_TOTAL},{R["sites"]}*{R["rent"]}*12)',
       CUR2, "auto: site mix or sites x rent", key="gpr")
u_calc(22, "Less: Vacancy & Credit Loss", f'=-{R["gpr"]}*{R["vac"]}', CUR2, key="vacloss")
u_calc(23, "Plus: Other Income", f'={R["other_m"]}*12', CUR2, key="other_a")
u_calc(24, "Effective Gross Income (EGI)", f'={R["gpr"]}+{R["vacloss"]}+{R["other_a"]}',
       CUR2, bold=True, fill=TOTAL_FILL, key="egi")

sec(ws, 26, "OPERATING EXPENSES  (Year 1)")
u_input(27, "Property Taxes", 18000, CUR2, key="tax")
u_input(28, "Insurance", 9000, CUR2, key="ins")
u_input(29, "Utilities (water/sewer/trash/elec)", 36000, CUR2, "owner-paid portion", key="util")
u_input(30, "Repairs & Maintenance", 15000, CUR2, key="rm")
u_input(31, "Payroll / Onsite Manager", 24000, CUR2, key="pay")
u_input(32, "Marketing, Admin, Office", 8000, CUR2, key="admin")
u_calc (33, "Management Fee", f'={R["mgmt_pct"]}*{R["egi"]}', CUR2, "= mgmt % x EGI", key="mgmt")
u_calc (34, "Replacement Reserves", f'={R["res_site"]}*{R["sites"]}', CUR2, key="reserves")
u_calc (35, "Total Operating Expenses", "=SUM(B27:B34)", CUR2, bold=True, fill=TOTAL_FILL, key="opex")
u_calc (36, "Operating Expense Ratio", f'={R["opex"]}/{R["egi"]}', PCT, "OpEx / EGI", key="oer")
u_calc (37, "Expenses per Site", f'={R["opex"]}/{R["sites"]}', CUR2)

sec(ws, 39, "NET OPERATING INCOME")
u_calc(40, "Net Operating Income (NOI)", f'={R["egi"]}-{R["opex"]}', CUR2,
       bold=True, fill=KEY_FILL, key="noi")
u_calc(41, "NOI per Site", f'={R["noi"]}/{R["sites"]}', CUR2)

sec(ws, 43, "FINANCING & EXIT ASSUMPTIONS")
u_input(44, "Loan-to-Value (LTV) %", 0.70, PCT, key="ltv")
u_input(45, "Interest Rate %", 0.07, PCT2, "annual", key="intr")
u_input(46, "Amortization (years)", 25, CUR, key="amort")
u_input(47, "Closing Costs %", 0.03, PCT, "of purchase price", key="close_pct")
u_input(48, "Initial CapEx / Rehab", 50000, CUR2, "day-one improvements", key="capex")
u_input(49, "Hold Period (years)", 5, CUR, "1-10", key="hold")
u_input(50, "Exit Cap Rate %", 0.085, PCT2, "for sale value", key="exitcap")
u_input(51, "Selling Costs %", 0.05, PCT, "of sale price", key="sellcost")

sec(ws, 53, "DEBT & EQUITY")
u_calc(54, "Loan Amount", f'={R["price"]}*{R["ltv"]}', CUR2, key="loan")
u_calc(55, "Down Payment (Equity)", f'={R["price"]}-{R["loan"]}', CUR2, key="down")
u_calc(56, "Closing Costs", f'={R["price"]}*{R["close_pct"]}', CUR2, key="close")
u_calc(57, "Total Cash Required", f'={R["down"]}+{R["close"]}+{R["capex"]}',
       CUR2, bold=True, fill=TOTAL_FILL, key="cash")
u_calc(58, "Monthly Debt Payment", f'=PMT({R["intr"]}/12,{R["amort"]}*12,-{R["loan"]})', CUR2, key="pmt_m")
u_calc(59, "Annual Debt Service", f'={R["pmt_m"]}*12', CUR2, bold=True, key="ads")

sec(ws, 61, "KEY RETURN METRICS  (Year 1)")
u_calc(62, "Going-in Cap Rate", f'={R["noi"]}/{R["price"]}', PCT, "NOI / Price",
       bold=True, fill=KEY_FILL, key="caprate")
u_calc(63, "Cash Flow Before Tax (CFBT)", f'={R["noi"]}-{R["ads"]}', CUR2,
       bold=True, fill=KEY_FILL, key="cfbt")
u_calc(64, "Cash-on-Cash Return", f'={R["cfbt"]}/{R["cash"]}', PCT, "CFBT / cash invested",
       bold=True, fill=KEY_FILL, key="coc")
u_calc(65, "Debt Service Coverage (DSCR)", f'={R["noi"]}/{R["ads"]}', MULT,
       "lenders want >= 1.25x", bold=True, fill=KEY_FILL, key="dscr")
u_calc(66, "Gross Rent Multiplier (GRM)", f'={R["price"]}/{R["gpr"]}', '0.00"x"')
u_calc(67, "Break-even Occupancy", f'=({R["opex"]}+{R["ads"]})/({R["gpr"]}+{R["other_a"]})', PCT,
       "occupancy to cover costs", key="breakeven")
u_calc(68, "Debt Yield", f'={R["noi"]}/{R["loan"]}', PCT, "NOI / loan (lender metric)")

# =============================================================================
# SHEET — ACTUALS vs PRO FORMA
# =============================================================================
av = wb.create_sheet("Actuals vs Pro Forma")
av.sheet_view.showGridLines = False
av.column_dimensions["A"].width = 36
for col in "BCD": av.column_dimensions[col].width = 18
merge_title(av, "A1:D1", "IN-PLACE ACTUALS  vs  STABILIZED PRO FORMA")
av["A2"] = ("Left column = what the seller is doing TODAY (from their T-12 P&L). Right = your stabilized plan "
            "(pulled from Underwriting). The gap is your value-add.")
av["A2"].font = NOTE
U = "Underwriting!"
av["A4"]="Line Item"; av["B4"]="In-Place / Actuals"; av["C4"]="Stabilized Pro Forma"; av["D4"]="Variance"
for cl in "ABCD": av[f"{cl}4"].font=SECTION; av[f"{cl}4"].fill=SECTION_FILL; av[f"{cl}4"].alignment=Alignment(horizontal="center")

def av_row(row, label, actual_val, stab_formula, fmt=CUR2, isinput=True, bold=False, fill=None):
    av[f"A{row}"]=label; av[f"A{row}"].font=BOLD if bold else LABEL
    a=av[f"B{row}"]
    if isinput:
        a.value=actual_val; a.font=INPUTF; a.fill=INPUT_FILL; a.border=BORDER
    else:
        a.value=actual_val; a.font=BOLD if bold else LABEL
        if fill: a.fill=fill
    a.number_format=fmt; a.alignment=Alignment(horizontal="right")
    c=av[f"C{row}"]; c.value=stab_formula; c.font=BOLD if bold else LABEL
    c.number_format=fmt; c.alignment=Alignment(horizontal="right")
    if fill: c.fill=fill
    d=av[f"D{row}"]
    if fmt!=PCT and fmt!=MULT:
        d.value=f"=C{row}-B{row}"; d.number_format=CUR2
    d.font=LABEL; d.alignment=Alignment(horizontal="right")

av_row(5,  "Gross Potential Rent", 240000, f"={U}{R['gpr']}")
av_row(6,  "Vacancy & Credit Loss", -48000, f"={U}{R['vacloss']}")
av_row(7,  "Other Income", 12000, f"={U}{R['other_a']}")
av_row(8,  "Effective Gross Income", "=B5+B6+B7", f"={U}{R['egi']}", isinput=False, bold=True, fill=TOTAL_FILL)
av_row(9,  "Property Taxes", 12000, f"={U}{R['tax']}")
av_row(10, "Insurance", 8000, f"={U}{R['ins']}")
av_row(11, "Utilities", 38000, f"={U}{R['util']}")
av_row(12, "Repairs & Maintenance", 20000, f"={U}{R['rm']}")
av_row(13, "Payroll / Onsite Mgr", 20000, f"={U}{R['pay']}")
av_row(14, "Marketing, Admin, Office", 6000, f"={U}{R['admin']}")
av_row(15, "Management Fee", 0, f"={U}{R['mgmt']}")
av_row(16, "Replacement Reserves", 0, f"={U}{R['reserves']}")
av_row(17, "Total Operating Expenses", "=SUM(B9:B16)", f"={U}{R['opex']}", isinput=False, bold=True, fill=TOTAL_FILL)
av_row(18, "Net Operating Income", "=B8-B17", f"={U}{R['noi']}", isinput=False, bold=True, fill=KEY_FILL)
av["A20"]="Cap Rate at Purchase Price"; av["A20"].font=BOLD
av["B20"]=f"=B18/{U}{R['price']}"; av["B20"].number_format=PCT; av["B20"].font=BOLD; av["B20"].fill=KEY_FILL; av["B20"].alignment=Alignment(horizontal="right")
av["C20"]=f"={U}{R['caprate']}"; av["C20"].number_format=PCT; av["C20"].font=BOLD; av["C20"].fill=KEY_FILL; av["C20"].alignment=Alignment(horizontal="right")
av["A22"]="In-place cap rate = what you're actually buying today. Underwrite to this first, the pro forma second."
av["A22"].font=NOTE

# =============================================================================
# SHEET — PRO FORMA & RETURNS
# =============================================================================
pf = wb.create_sheet("Pro Forma & Returns")
pf.sheet_view.showGridLines = False
pf.column_dimensions["A"].width = 34
for i in range(11):
    pf.column_dimensions[get_column_letter(2+i)].width = 13
merge_title(pf, "A1:L1", "10-YEAR PRO FORMA & RETURNS")
def uref(k): return U + R[k]

pf["A3"]="Year"; pf["A3"].font=SECTION; pf["A3"].fill=SECTION_FILL
for y in range(0,11):
    col=get_column_letter(2+y); cc=pf[f"{col}3"]; cc.value=y
    cc.font=SECTION; cc.fill=SECTION_FILL; cc.alignment=Alignment(horizontal="center")

def pf_row(row, label, fn, fmt=CUR2, bold=False, fill=None):
    pf[f"A{row}"]=label; pf[f"A{row}"].font=BOLD if bold else LABEL
    for y in range(0,11):
        col=get_column_letter(2+y); v=fn(y,col)
        if v is None: continue
        cell=pf[f"{col}{row}"]; cell.value=v; cell.number_format=fmt
        cell.font=BOLD if bold else LABEL; cell.alignment=Alignment(horizontal="right")
        if fill: cell.fill=fill

pf_row(5, "Gross Potential Rent",
       lambda y,c: None if y==0 else f"={uref('gpr')}*(1+{uref('rent_g')})^{y-1}")
pf_row(6, "Less: Vacancy & Credit Loss",
       lambda y,c: None if y==0 else f"=-{c}5*{uref('vac')}")
pf_row(7, "Other Income",
       lambda y,c: None if y==0 else f"={uref('other_a')}*(1+{uref('rent_g')})^{y-1}")
pf_row(8, "Effective Gross Income",
       lambda y,c: None if y==0 else f"={c}5+{c}6+{c}7", bold=True, fill=TOTAL_FILL)
pf_row(9, "Operating Expenses",
       lambda y,c: None if y==0 else f"={uref('opex')}*(1+{uref('exp_g')})^{y-1}")
pf_row(10, "Net Operating Income",
       lambda y,c: None if y==0 else f"={c}8-{c}9", bold=True, fill=KEY_FILL)
pf_row(11, "Annual Debt Service",
       lambda y,c: None if y==0 else f"=IF({y}<={uref('hold')},{uref('ads')},0)")
def bal_y(y,c):
    if y==0: return f"={uref('loan')}"
    i=f"({uref('intr')}/12)"; k=f"MIN({y},{uref('amort')})*12"
    return f"={uref('loan')}*(1+{i})^({k})-{uref('pmt_m')}*(((1+{i})^({k})-1)/{i})"
pf_row(12, "Loan Balance (end of yr)", bal_y)
pf_row(13, "Operating Cash Flow",
       lambda y,c: None if y==0 else f"=IF({y}<={uref('hold')},{c}10-{c}11,0)", bold=True)
def sale_y(y,c):
    if y==0: return None
    saleprice=f"({c}10*(1+{uref('rent_g')}))/{uref('exitcap')}"
    net=f"({saleprice})*(1-{uref('sellcost')})-{c}12"
    return f"=IF({y}={uref('hold')},{net},0)"
pf_row(14, "Net Sale Proceeds (reversion)", sale_y, fill=TOTAL_FILL)
pf_row(15, "Total Cash Flow (for IRR)",
       lambda y,c: f"=-{uref('cash')}" if y==0 else f"={c}13+{c}14",
       bold=True, fill=KEY_FILL)

sec(pf, 17, "RETURN SUMMARY (over hold period)", "A:D")
def summ(row, label, formula, fmt, note=""):
    pf[f"A{row}"]=label; pf[f"A{row}"].font=BOLD
    c=pf[f"B{row}"]; c.value=formula; c.number_format=fmt; c.font=BOLD
    c.fill=KEY_FILL; c.alignment=Alignment(horizontal="right")
    if note: pf[f"C{row}"]=note; pf[f"C{row}"].font=NOTE
summ(18, "Levered IRR", "=IRR(B15:L15)", PCT, "internal rate of return on equity")
summ(19, "Equity Multiple", "=(SUM(C13:L13)+SUM(C14:L14))/-B15", MULT, "total cash returned / equity")
summ(20, "Avg. Cash-on-Cash", f"=AVERAGEIF(C11:L11,\">0\",C13:L13)/{uref('cash')}", PCT,
     "avg operating CF / equity during hold")
summ(21, "Year-1 Cash-on-Cash", f"={uref('coc')}", PCT)
summ(22, "Going-in Cap Rate", f"={uref('caprate')}", PCT)
summ(23, "Total Profit", "=SUM(C13:L13)+SUM(C14:L14)+B15", CUR2, "all cash flows incl. equity out")

# =============================================================================
# SHEET — SENSITIVITY
# =============================================================================
sn = wb.create_sheet("Sensitivity")
sn.sheet_view.showGridLines = False
sn.column_dimensions["A"].width = 22
for col in "BCDEF": sn.column_dimensions[col].width = 13
merge_title(sn, "A1:F1", "SENSITIVITY TABLES")
sn["A2"]="Rows = purchase price scenarios. Columns = interest rate. Cells recompute live from Underwriting inputs."
sn["A2"].font=NOTE

price_factors=[0.90,0.95,1.00,1.05,1.10]
rate_deltas=[-0.01,-0.005,0,0.005,0.01]

def build_grid(top, title_text, metric, fmt):
    sec(sn, top, title_text, "A:F")
    # corner + column headers (interest rate)
    hdr=top+1
    sn[f"A{hdr}"]="Price \\ Rate"; sn[f"A{hdr}"].font=BOLD; sn[f"A{hdr}"].alignment=Alignment(horizontal="center")
    sn[f"A{hdr}"].fill=TOTAL_FILL
    for j,d in enumerate(rate_deltas):
        col=get_column_letter(2+j); c=sn[f"{col}{hdr}"]
        c.value=f"={uref('intr')}+{d}"; c.number_format=PCT2; c.font=BOLD
        c.fill=TOTAL_FILL; c.alignment=Alignment(horizontal="center")
    # rows
    for i,f in enumerate(price_factors):
        row=hdr+1+i; c=sn[f"A{row}"]
        c.value=f"={uref('price')}*{f}"; c.number_format=CUR2; c.font=BOLD
        c.fill=TOTAL_FILL; c.alignment=Alignment(horizontal="right")
        for j,d in enumerate(rate_deltas):
            col=get_column_letter(2+j); cell=sn[f"{col}{row}"]
            P=f"$A{row}"; r=f"{col}${hdr}"
            loan=f"({P}*{uref('ltv')})"
            pmt=f"({loan}*({r}/12)/(1-(1+{r}/12)^(-{uref('amort')}*12)))"
            ads=f"({pmt}*12)"
            cash=f"({P}*(1-{uref('ltv')})+{P}*{uref('close_pct')}+{uref('capex')})"
            if metric=="coc":
                cell.value=f"=({uref('noi')}-{ads})/{cash}"
            elif metric=="dscr":
                cell.value=f"={uref('noi')}/{ads}"
            cell.number_format=fmt; cell.alignment=Alignment(horizontal="right")
    return hdr+1+len(price_factors)

end1=build_grid(4, "CASH-ON-CASH RETURN  —  Price x Interest Rate", "coc", PCT)
end2=build_grid(end1+2, "DEBT SERVICE COVERAGE (DSCR)  —  Price x Interest Rate", "dscr", MULT)
sn[f"A{end2+1}"]="NOTE: the two tables above hold NOI constant to isolate price & financing risk."
sn[f"A{end2+1}"].font=NOTE

# ---- rent x vacancy grid (operational risk; recomputes NOI) ----------------
rent_deltas=[-100,-50,0,50,100]
vac_values=[0.05,0.10,0.15,0.20,0.25]
top=end2+3
sec(sn, top, "CASH-ON-CASH RETURN  —  Monthly Rent x Vacancy", "A:F")
hdr=top+1
sn[f"A{hdr}"]="Rent \\ Vacancy"; sn[f"A{hdr}"].font=BOLD; sn[f"A{hdr}"].fill=TOTAL_FILL
sn[f"A{hdr}"].alignment=Alignment(horizontal="center")
for j,v in enumerate(vac_values):
    col=get_column_letter(2+j); c=sn[f"{col}{hdr}"]
    c.value=v; c.number_format=PCT; c.font=BOLD; c.fill=TOTAL_FILL
    c.alignment=Alignment(horizontal="center")
for i,d in enumerate(rent_deltas):
    row=hdr+1+i; c=sn[f"A{row}"]
    c.value=f"={uref('rent')}+{d}"; c.number_format=CUR2; c.font=BOLD
    c.fill=TOTAL_FILL; c.alignment=Alignment(horizontal="right")
    for j,v in enumerate(vac_values):
        col=get_column_letter(2+j); cell=sn[f"{col}{row}"]
        rent=f"$A{row}"; vac=f"{col}${hdr}"
        egi=f"({uref('sites')}*{rent}*12*(1-{vac})+{uref('other_a')})"
        # opex excl. mgmt is fixed; mgmt scales with EGI
        opex=f"(({uref('opex')}-{uref('mgmt')})+{uref('mgmt_pct')}*{egi})"
        noi=f"({egi}-{opex})"
        cell.value=f"=({noi}-{uref('ads')})/{uref('cash')}"
        cell.number_format=PCT; cell.alignment=Alignment(horizontal="right")
endR=hdr+1+len(rent_deltas)
sn[f"A{endR+1}"]="NOTE: this table recomputes NOI (management fee scales with EGI). Uses sites x rent income, not the Income Detail build."
sn[f"A{endR+1}"].font=NOTE

# =============================================================================
# SHEET — DEAL SUMMARY (one-page, lender-friendly)
# =============================================================================
ds = wb.create_sheet("Deal Summary")
ds.sheet_view.showGridLines = False
ds.column_dimensions["A"].width = 30
ds.column_dimensions["B"].width = 22
ds.column_dimensions["C"].width = 30
ds.column_dimensions["D"].width = 22
merge_title(ds, "A1:D1", "DEAL SUMMARY")
PFR = "'Pro Forma & Returns'!"

ds.merge_cells("A2:D2")
ds["A2"]=f'=Underwriting!{R["name"]}&"  —  "&Underwriting!{R["loc"]}'
ds["A2"].font=Font(size=13, bold=True, color="1F4E78"); ds["A2"].alignment=Alignment(horizontal="center")

def kv(cell_label, cell_val, label, formula, fmt, big=False, fill=None):
    l=ds[cell_label]; l.value=label; l.font=BOLD
    v=ds[cell_val]; v.value=formula; v.number_format=fmt
    v.font=BIG if big else BOLD; v.alignment=Alignment(horizontal="right")
    if fill: v.fill=fill; l.fill=fill

sec(ds, 4, "THE ASSET", "A:D")
kv("A5","B5","Number of Sites", f"=Underwriting!{R['sites']}", CUR)
kv("C5","D5","Purchase Price", f"=Underwriting!{R['price']}", CUR2)
kv("A6","B6","Price per Site", f"=Underwriting!{R['price']}/Underwriting!{R['sites']}", CUR2)
kv("C6","D6","Year-1 NOI", f"=Underwriting!{R['noi']}", CUR2)

sec(ds, 8, "FINANCING", "A:D")
kv("A9","B9","Loan Amount", f"=Underwriting!{R['loan']}", CUR2)
kv("C9","D9","Down Payment (Equity)", f"=Underwriting!{R['down']}", CUR2)
kv("A10","B10","Interest Rate", f"=Underwriting!{R['intr']}", PCT2)
kv("C10","D10","Amortization (yrs)", f"=Underwriting!{R['amort']}", CUR)
kv("A11","B11","Annual Debt Service", f"=Underwriting!{R['ads']}", CUR2)
kv("C11","D11","Total Cash Required", f"=Underwriting!{R['cash']}", CUR2)

sec(ds, 13, "RETURNS — THE HEADLINE NUMBERS", "A:D")
kv("A14","B14","Going-in Cap Rate", f"=Underwriting!{R['caprate']}", PCT, big=True, fill=KEY_FILL)
kv("C14","D14","DSCR", f"=Underwriting!{R['dscr']}", MULT, big=True, fill=KEY_FILL)
kv("A16","B16","Cash-on-Cash (Yr 1)", f"=Underwriting!{R['coc']}", PCT, big=True, fill=KEY_FILL)
kv("C16","D16","Debt Yield", f"=Underwriting!{R['noi']}/Underwriting!{R['loan']}", PCT, big=True, fill=KEY_FILL)
kv("A18","B18","Levered IRR", f"={PFR}B18", PCT, big=True, fill=KEY_FILL)
kv("C18","D18","Equity Multiple", f"={PFR}B19", MULT, big=True, fill=KEY_FILL)

sec(ds, 20, "EXIT ASSUMPTIONS", "A:D")
kv("A21","B21","Hold Period (yrs)", f"=Underwriting!{R['hold']}", CUR)
kv("C21","D21","Exit Cap Rate", f"=Underwriting!{R['exitcap']}", PCT2)
kv("A22","B22","Annual Rent Growth", f"=Underwriting!{R['rent_g']}", PCT)
kv("C22","D22","Annual Expense Growth", f"=Underwriting!{R['exp_g']}", PCT)

ds.merge_cells("A24:D24")
ds["A24"]="Screening tool only — not investment advice. Verify every figure against the seller's T-12 and rent roll."
ds["A24"].font=NOTE

# =============================================================================
# SHEET — LOAN SIZING (max purchase price at target DSCR / debt yield)
# =============================================================================
ls = wb.create_sheet("Loan Sizing")
ls.sheet_view.showGridLines=False
ls.column_dimensions["A"].width=40
ls.column_dimensions["B"].width=18
ls.column_dimensions["C"].width=34
merge_title(ls, "A1:C1", "LOAN SIZING — MAX PRICE AT TARGET DSCR")
ls["A2"]=("Works backward from how much debt the NOI supports. NOI is held at the Underwriting Year-1 figure; "
          "rate, amortization and LTV come from the Underwriting tab.")
ls["A2"].font=NOTE

def ls_input(row, label, val, fmt, note="", key=None):
    ls[f"A{row}"]=label; ls[f"A{row}"].font=LABEL
    c=ls[f"B{row}"]; c.value=val; c.font=INPUTF; c.fill=INPUT_FILL; c.border=BORDER
    c.number_format=fmt; c.alignment=Alignment(horizontal="right")
    if note: ls[f"C{row}"]=note; ls[f"C{row}"].font=NOTE
    if key: LS[key]=f"B{row}"
def ls_calc(row, label, formula, fmt, note="", bold=False, fill=None, key=None):
    ls[f"A{row}"]=label; ls[f"A{row}"].font=BOLD if bold else LABEL
    c=ls[f"B{row}"]; c.value=formula; c.font=BOLD if bold else LABEL
    c.number_format=fmt; c.alignment=Alignment(horizontal="right")
    if fill: c.fill=fill
    if note: ls[f"C{row}"]=note; ls[f"C{row}"].font=NOTE
    if key: LS[key]=f"B{row}"
LS={}

sec(ls, 4, "TARGETS  (your / the lender's limits)", "A:C")
ls_input(5, "Target DSCR", 1.25, MULT, "minimum coverage lender requires", key="tdscr")
ls_input(6, "Target Debt Yield", 0.10, PCT, "NOI / loan floor (often 9-10%)", key="tdy")
ls_input(7, "Max LTV %", 0.75, PCT, "lender's loan-to-value cap", key="maxltv")

sec(ls, 9, "LOAN INPUTS  (from Underwriting)", "A:C")
ls_calc(10, "Year-1 NOI", f"=Underwriting!{R['noi']}", CUR2, key="noi")
ls_calc(11, "Interest Rate", f"=Underwriting!{R['intr']}", PCT2, key="rate")
ls_calc(12, "Amortization (yrs)", f"=Underwriting!{R['amort']}", CUR, key="amort")
ls_calc(13, "Annual Debt Constant (per $1 loan)",
        f"=PMT({LS['rate']}/12,{LS['amort']}*12,-1)*12", '0.0000',
        "= annual debt service per $1 borrowed", key="k")

sec(ls, 15, "MAXIMUM SUPPORTABLE LOAN", "A:C")
ls_calc(16, "Max Annual Debt Service (DSCR)", f"={LS['noi']}/{LS['tdscr']}", CUR2,
        "= NOI / target DSCR", key="maxads")
ls_calc(17, "Max Loan — DSCR constrained", f"={LS['maxads']}/{LS['k']}", CUR2, key="loan_dscr")
ls_calc(18, "Max Loan — Debt-Yield constrained", f"={LS['noi']}/{LS['tdy']}", CUR2, key="loan_dy")
ls_calc(19, "Binding Max Loan", f"=MIN({LS['loan_dscr']},{LS['loan_dy']})", CUR2,
        "the lesser of the two", bold=True, fill=TOTAL_FILL, key="loan_max")

sec(ls, 21, "MAXIMUM PURCHASE PRICE", "A:C")
ls_calc(22, "Max Price — DSCR/Debt-Yield (at Max LTV)", f"={LS['loan_max']}/{LS['maxltv']}", CUR2,
        "price the debt supports at your LTV", bold=True, fill=KEY_FILL, key="price_debt")
ls_calc(23, "Current Purchase Price", f"=Underwriting!{R['price']}", CUR2, key="price_now")
ls_calc(24, "Price Headroom ($)", f"={LS['price_debt']}-{LS['price_now']}", CUR2,
        "positive = room to pay more; negative = overpaying", bold=True, key="headroom")
ls_calc(25, "Price Headroom (%)", f"={LS['headroom']}/{LS['price_now']}", PCT, bold=True)
ls_calc(26, "DSCR at Current Price", f"=Underwriting!{R['dscr']}", MULT,
        "for reference (target above)")
ls_calc(27, "Required Equity at Max Price", f"={LS['price_debt']}*(1-{LS['maxltv']})", CUR2,
        "down payment if you buy at the max")

ls["A29"]=("READ: if Headroom is negative, the price is higher than the income can safely finance at your target "
           "DSCR — you'd need more equity, a lower rate, or a price cut.")
ls["A29"].font=NOTE

# =============================================================================
# SHEET — NORMALIZATION (Linda's "build your own NOI")
# =============================================================================
nm = wb.create_sheet("Normalization")
nm.sheet_view.showGridLines=False
for cl,w in (("A",34),("B",16),("C",16),("D",42)):
    nm.column_dimensions[cl].width=w
merge_title(nm, "A1:D1", "NORMALIZE THE NOI — BUILD YOUR OWN NUMBERS")
nm["A2"]=("Linda: the T-12 is your starting point but never trusted as-is. Enter the SELLER'S reported figures "
          "(yellow); the Normalized column rebuilds the NOI with Linda's 4 adjustments. Part of the LindaAI framework.")
nm["A2"].font=NOTE

sec(nm, 4, "LINDA'S NORMALIZATION RULES (editable)", "A:D")
def nm_in(row,label,val,fmt,note):
    nm[f"A{row}"]=label; nm[f"A{row}"].font=LABEL
    c=nm[f"B{row}"]; c.value=val; c.number_format=fmt; c.font=INPUTF; c.fill=INPUT_FILL
    c.border=BORDER; c.alignment=Alignment(horizontal="right")
    if note: nm[f"D{row}"]=note; nm[f"D{row}"].font=NOTE
nm_in(5,"Management Fee (% of EGI)",0.10,PCT,"self-managed owners show little or none")
nm_in(6,"Repairs & Maintenance (% of EGI)",0.05,PCT,"or use the prior-year actual")
nm_in(7,"CapEx / Reserves (% of EGI)",0.03,PCT,"a real cost that's rarely in the P&L")
nm_in(8,"Property Tax bump (%)",0.20,PCT,"taxes reset on sale in disclosure states")
MG,RM,CX,TX="B5","B6","B7","B8"

sec(nm, 10, "INCOME", "A:D")
nm["A11"]="Effective Gross Income (EGI)"; nm["A11"].font=LABEL
c=nm["B11"]; c.value=247500; c.number_format=CUR2; c.font=INPUTF; c.fill=INPUT_FILL
c.border=BORDER; c.alignment=Alignment(horizontal="right")
nm["C11"]="=B11"; nm["C11"].number_format=CUR2; nm["C11"].font=BOLD; nm["C11"].alignment=Alignment(horizontal="right")
nm["D11"]="seller's real top line (Linda: gross income = EGI)"; nm["D11"].font=NOTE
EGI="B11"

sec(nm, 13, "OPERATING EXPENSES", "A:D")
nm["B14"]="Seller Reported"; nm["C14"]="Normalized"; nm["D14"]="Linda adjustment"
for cl in ("B","C","D"):
    nm[f"{cl}14"].font=SECTION; nm[f"{cl}14"].fill=SECTION_FILL; nm[f"{cl}14"].alignment=Alignment(horizontal="center")
# (label, seller_default, normalized_formula_template, note)  {r}=current row
exp=[
 ("Property Taxes", 12000, "=B{r}*(1+%s)"%TX, "+20%: resets on sale"),
 ("Insurance", 8000, "=B{r}", "pass-through"),
 ("Utilities (W/S/trash/elec)", 38000, "=B{r}", "pass-through"),
 ("Repairs & Maintenance", 4200, "=MAX(B{r},%s*%s)"%(RM,EGI), "use >=5%% of EGI (deferred maint)"),
 ("Management Fee", 3500, "=MAX(B{r},%s*%s)"%(MG,EGI), "use >=10%% of EGI (self-mgmt understates)"),
 ("CapEx / Reserves", 0, "=MAX(B{r},%s*%s)"%(CX,EGI), "add >=3%% of EGI (never in P&L)"),
 ("Payroll / Onsite", 0, "=B{r}", "pass-through"),
 ("Marketing / Admin / Other", 6000, "=B{r}", "pass-through"),
]
er=15
efirst=er
for label,sv,nf,note in exp:
    nm[f"A{er}"]=label; nm[f"A{er}"].font=LABEL
    b=nm[f"B{er}"]; b.value=sv; b.number_format=CUR2; b.font=INPUTF; b.fill=INPUT_FILL
    b.border=BORDER; b.alignment=Alignment(horizontal="right")
    cc=nm[f"C{er}"]; cc.value=nf.format(r=er); cc.number_format=CUR2; cc.alignment=Alignment(horizontal="right")
    nm[f"D{er}"]=note; nm[f"D{er}"].font=NOTE
    er+=1
elast=er-1
# totals
nm[f"A{er}"]="Total Operating Expenses"; nm[f"A{er}"].font=BOLD
nm[f"B{er}"]=f"=SUM(B{efirst}:B{elast})"; nm[f"C{er}"]=f"=SUM(C{efirst}:C{elast})"
for col in ("B","C"):
    nm[f"{col}{er}"].number_format=CUR2; nm[f"{col}{er}"].font=BOLD; nm[f"{col}{er}"].fill=TOTAL_FILL
    nm[f"{col}{er}"].alignment=Alignment(horizontal="right")
opex_s,opex_n=f"B{er}",f"C{er}"
rr=er+1
nm[f"A{rr}"]="Expense Ratio"; nm[f"A{rr}"].font=LABEL
nm[f"B{rr}"]=f"={opex_s}/{EGI}"; nm[f"C{rr}"]=f"={opex_n}/C11"
for col in ("B","C"):
    nm[f"{col}{rr}"].number_format=PCT; nm[f"{col}{rr}"].alignment=Alignment(horizontal="right")
nm[f"D{rr}"]="Linda: under 30% means something's missing"; nm[f"D{rr}"].font=NOTE

sec(nm, rr+2, "RESULT", "A:D")
nr=rr+3
nm[f"A{nr}"]="Net Operating Income (NOI)"; nm[f"A{nr}"].font=BOLD
nm[f"B{nr}"]=f"={EGI}-{opex_s}"; nm[f"C{nr}"]=f"=C11-{opex_n}"
for col in ("B","C"):
    nm[f"{col}{nr}"].number_format=CUR2; nm[f"{col}{nr}"].font=BOLD; nm[f"{col}{nr}"].fill=KEY_FILL
    nm[f"{col}{nr}"].alignment=Alignment(horizontal="right")
noi_s,noi_n=f"B{nr}",f"C{nr}"
nm[f"A{nr+1}"]="NOI Haircut ($ cut from seller's NOI)"; nm[f"A{nr+1}"].font=BOLD
nm[f"C{nr+1}"]=f"={noi_s}-{noi_n}"; nm[f"C{nr+1}"].number_format=CUR2; nm[f"C{nr+1}"].font=BOLD
nm[f"C{nr+1}"].alignment=Alignment(horizontal="right")
nm[f"A{nr+2}"]="NOI Haircut (%)"; nm[f"A{nr+2}"].font=LABEL
nm[f"C{nr+2}"]=f"=({noi_s}-{noi_n})/{noi_s}"; nm[f"C{nr+2}"].number_format=PCT; nm[f"C{nr+2}"].alignment=Alignment(horizontal="right")
nm[f"A{nr+3}"]="Valuation Cap Rate (for overpayment calc)"; nm[f"A{nr+3}"].font=LABEL
capcell=nm[f"B{nr+3}"]; capcell.value=0.10; capcell.number_format=PCT2; capcell.font=INPUTF
capcell.fill=INPUT_FILL; capcell.border=BORDER; capcell.alignment=Alignment(horizontal="right")
nm[f"A{nr+4}"]="OVERPAYMENT RISK  (haircut / cap rate)"; nm[f"A{nr+4}"].font=Font(size=12,bold=True,color="9C0006")
ovp=nm[f"C{nr+4}"]; ovp.value=f"=({noi_s}-{noi_n})/B{nr+3}"; ovp.number_format=CUR2
ovp.font=Font(size=12,bold=True,color="9C0006"); ovp.fill=WARN_FILL; ovp.alignment=Alignment(horizontal="right")
nm[f"D{nr+4}"]="Linda: a small NOI cut = a huge price difference"; nm[f"D{nr+4}"].font=NOTE
nm[f"A{nr+6}"]=("NEXT: carry the NORMALIZED expense figures (column C) into the Underwriting tab so every "
                "metric and the Deal Scorecard reflect YOUR numbers, not the seller's.")
nm[f"A{nr+6}"].font=NOTE

# =============================================================================
# SHEET — DEAL SCORECARD (GOOD / OK / BAD rating)
# =============================================================================
sc = wb.create_sheet("Deal Scorecard")
sc.sheet_view.showGridLines=False
for cl,w in (("A",36),("B",13),("C",27),("D",9),("E",9),("F",9),("G",9),("H",16)):
    sc.column_dimensions[cl].width=w
merge_title(sc, "A1:H1", "DEAL SCORECARD — LINDAAI FRAMEWORK")
sc["A2"]=("Linda's framework. Underwrite to the NORMALIZED NOI (see Normalization tab) before trusting these. "
          "GREEN cutoff cells are editable. Part of the LindaAI underwriting framework.")
sc["A2"].font=NOTE

GREEN=PatternFill("solid", fgColor="C6EFCE")
YELL =PatternFill("solid", fgColor="FFEB9C")
RED  =PatternFill("solid", fgColor="FFC7CE")
GFONT=Font(color="006100", bold=True)
YFONT=Font(color="9C6500", bold=True)
RFONT=Font(color="9C0006", bold=True)

hdr=4
heads=["Metric","Deal Value","Linda's rule","Good","OK","(band)","(band)","Rating"]
for j,h in enumerate(heads):
    col=get_column_letter(1+j); c=sc[f"{col}{hdr}"]
    c.value=h; c.font=SECTION; c.fill=SECTION_FILL; c.alignment=Alignment(horizontal="center")

U_="Underwriting!"
egiv=U_+R['egi']; price=U_+R['price']; intr=U_+R['intr']; noi=U_+R['noi']
dscrv=U_+R['dscr']; cocv=U_+R['coc']; oerv=U_+R['oer']; caprate=U_+R['caprate']

# rows: ("__sec__",title) OR (label, value_formula, fmt, kind, cuts, rule)
# kind: high|low (cuts=good,ok)  ;  band (cuts=badlo,goodlo,goodhi,badhi)
rows=[
 ("__sec__","GO / NO-GO METRICS  (after normalizing the NOI)"),
 ("DSCR", f"={dscrv}", MULT, "high", (1.50,1.35), "Linda floor: 1.35x to GO"),
 ("Cash-on-Cash Return", f"={cocv}", PCT, "high", (0.10,0.08), "Linda target: 10%+ to GO"),
 ("__sec__","THE 5 QUICK HACKS  (screen any listing in ~2 min)"),
 ("1% Rule  (monthly gross / price)", f"=({egiv}/12)/{price}", PCT2, "high", (0.01,0.009), ">=1% likely good"),
 ("10x Rule — GRM  (price / annual gross)", f"={price}/{egiv}", '0.00"x"', "low", (10,12), "<10 ideal; >12 pricey"),
 ("Quick DSCR  (NOI / (price x rate))", f"={noi}/({price}*{intr})", MULT, "high", (1.5,1.25), ">1.5 good; <1.25 risk"),
 ("Expense Ratio  (OpEx / gross)", f"={oerv}", PCT, "band", (0.30,0.35,0.65,0.70), "35-65% healthy; <30%=fiction"),
 ("Leverage spread  (cap - interest)", f"={caprate}-{intr}", PCT, "high", (0.02,0.0), "cap>rate = positive leverage"),
]
r=hdr+1
row_dscr=row_coc=None
for item in rows:
    if item[0]=="__sec__":
        sec(sc, r, item[1], "A:H"); r+=1; continue
    label,vf,fmt,kind,cuts,rule=item
    sc[f"A{r}"]=label; sc[f"A{r}"].font=LABEL
    b=sc[f"B{r}"]; b.value=vf; b.number_format=fmt; b.font=BOLD; b.alignment=Alignment(horizontal="right")
    sc[f"C{r}"]=rule; sc[f"C{r}"].font=NOTE
    if kind=="band":
        for colcell,val in zip(("D","E","F","G"),cuts):
            c=sc[f"{colcell}{r}"]; c.value=val; c.number_format=fmt
            c.font=INPUTF; c.fill=INPUT_FILL; c.border=BORDER; c.alignment=Alignment(horizontal="right")
        rf=f'=IF(OR(B{r}<D{r},B{r}>G{r}),"BAD",IF(AND(B{r}>=E{r},B{r}<=F{r}),"GOOD","OK"))'
    else:
        for colcell,val in (("D",cuts[0]),("E",cuts[1])):
            c=sc[f"{colcell}{r}"]; c.value=val; c.number_format=fmt
            c.font=INPUTF; c.fill=INPUT_FILL; c.border=BORDER; c.alignment=Alignment(horizontal="right")
        if kind=="high":
            rf=f'=IF(B{r}>=D{r},"GOOD",IF(B{r}>=E{r},"OK","BAD"))'
        else:
            rf=f'=IF(B{r}<=D{r},"GOOD",IF(B{r}<=E{r},"OK","BAD"))'
    h=sc[f"H{r}"]; h.value=rf; h.alignment=Alignment(horizontal="center"); h.font=BOLD
    if label=="DSCR": row_dscr=r
    if label=="Cash-on-Cash Return": row_coc=r
    r+=1
last=r-1

rng=f"H{hdr+1}:H{last}"
sc.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"GOOD"'], fill=GREEN, font=GFONT))
sc.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"OK"'],   fill=YELL,  font=YFONT))
sc.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"BAD"'],  fill=RED,   font=RFONT))

sc[f"A{last+1}"]=("Expense Ratio is a BAND: below D = BAD (NOI likely inflated / 'fiction') · D-E and F-G = OK · "
                  "E-F = GOOD (healthy) · above G = BAD.  D,E,F,G default to 30/35/65/70%.")
sc[f"A{last+1}"].font=NOTE

# ---- Linda verdict: BOTH core metrics must clear --------------------------------
vr=last+3
sec(sc, vr-1, "VERDICT  (Linda's rule: both DSCR and Cash-on-Cash must clear)", "A:H")
sc[f"A{vr}"]="DSCR clears floor?"; sc[f"A{vr}"].font=BOLD
sc[f"B{vr}"]=f'=IF(B{row_dscr}>=E{row_dscr},"YES","NO")'; sc[f"B{vr}"].font=BOLD; sc[f"B{vr}"].alignment=Alignment(horizontal="center")
sc[f"A{vr+1}"]="Cash-on-Cash clears target?"; sc[f"A{vr+1}"].font=BOLD
sc[f"B{vr+1}"]=f'=IF(B{row_coc}>=D{row_coc},"YES","NO")'; sc[f"B{vr+1}"].font=BOLD; sc[f"B{vr+1}"].alignment=Alignment(horizontal="center")
sc[f"A{vr+3}"]="VERDICT"; sc[f"A{vr+3}"].font=Font(size=13,bold=True)
v=sc[f"B{vr+3}"]
v.value=(f'=IF(AND(B{row_dscr}>=E{row_dscr},B{row_coc}>=D{row_coc}),"GO - both clear",'
         f'IF(OR(B{row_dscr}>=E{row_dscr},B{row_coc}>=D{row_coc}),"CONDITIONAL - restructure","LIKELY NO-GO"))')
v.font=Font(size=13,bold=True); v.alignment=Alignment(horizontal="center")
sc.merge_cells(f"B{vr+3}:E{vr+3}")
vcell=f"$B${vr+3}"
sc.conditional_formatting.add(f"B{vr+3}", FormulaRule(formula=[f'LEFT({vcell},2)="GO"'], fill=GREEN, font=GFONT))
sc.conditional_formatting.add(f"B{vr+3}", FormulaRule(formula=[f'ISNUMBER(SEARCH("CONDITIONAL",{vcell}))'], fill=YELL, font=YFONT))
sc.conditional_formatting.add(f"B{vr+3}", FormulaRule(formula=[f'ISNUMBER(SEARCH("NO-GO",{vcell}))'], fill=RED, font=RFONT))
sc[f"A{vr+5}"]="Red flags (BAD ratings)"; sc[f"A{vr+5}"].font=BOLD
sc[f"B{vr+5}"]=f'=COUNTIF(H{hdr+1}:H{last},"BAD")'; sc[f"B{vr+5}"].font=BOLD; sc[f"B{vr+5}"].alignment=Alignment(horizontal="center")
sc[f"A{vr+7}"]=("CONDITIONAL means restructure to save it (Linda's 3 levers): (1) lower price to your MAO, "
                "(2) ask the seller to carry a slice — blended rate drops, DSCR clears, CoC jumps, "
                "(3) more down payment. If none work and the seller won't budge, your underwriting just "
                "saved you from an expensive mistake.")
sc[f"A{vr+7}"].font=NOTE

# =============================================================================
# SHEET — OFFER STRUCTURES (Linda's 4-column engine + MAO + seller carry)
# =============================================================================
of = wb.create_sheet("Offer Structures")
of.sheet_view.showGridLines=False
of.column_dimensions["A"].width=34
for cl in "BCDE": of.column_dimensions[cl].width=17
merge_title(of, "A1:E1", "OFFER STRUCTURES — MAO & SELLER CARRY (LINDA FRAMEWORK)")
of["A2"]=("NOI is held constant (operations don't change with financing). Each column is a price + financing "
          "scenario; yellow = editable levers. Verdict uses Linda's rule: DSCR >= target AND Cash-on-Cash >= target.")
of["A2"].font=NOTE

Uno=f"Underwriting!{R['noi']}"; Uclose=f"Underwriting!{R['close_pct']}"; Ucapex=f"Underwriting!{R['capex']}"

# ---- MAO auto-solve ---------------------------------------------------------
sec(of, 4, "MAXIMUM ALLOWABLE OFFER — auto-solved (conventional financing)", "A:E")
MAO={}
def of_kv(row,label,val,fmt,note="",inp=False,key=None,bold=False,fill=None,big=False):
    of[f"A{row}"]=label; of[f"A{row}"].font=Font(size=12,bold=True) if big else (BOLD if bold else LABEL)
    c=of[f"B{row}"]; c.value=val; c.number_format=fmt
    if inp: c.font=INPUTF; c.fill=INPUT_FILL; c.border=BORDER
    else: c.font=Font(size=12,bold=True) if big else (BOLD if bold else LABEL)
    if fill: c.fill=fill
    c.alignment=Alignment(horizontal="right")
    if note: of[f"C{row}"]=note; of[f"C{row}"].font=NOTE; of.merge_cells(f"C{row}:E{row}")
    if key: MAO[key]=f"B{row}"
of_kv(5,"Target DSCR",1.35,MULT,"Linda floor",inp=True,key="tdscr")
of_kv(6,"Target Cash-on-Cash",0.10,PCT,"Linda target",inp=True,key="tcoc")
of_kv(7,"Bank LTV % (conventional)",0.70,PCT,"loan-to-value for the conventional offer",inp=True,key="ltv")
of_kv(8,"Bank Interest Rate",0.07,PCT2,inp=True,key="rate")
of_kv(9,"Amortization (yrs)",25,CUR,inp=True,key="amort")
of_kv(10,"NOI (from Underwriting)",f"={Uno}",CUR2,key="noi")
of_kv(11,"Annual Debt Constant k (per $1 loan)",f"=PMT({MAO['rate']}/12,{MAO['amort']}*12,-1)*12",'0.0000',
      "annual debt service per $1 borrowed",key="k")
of_kv(12,"Max Price @ DSCR target",f"={MAO['noi']}/({MAO['tdscr']}*{MAO['ltv']}*{MAO['k']})",CUR2,
      "DSCR-constrained ceiling",key="pdscr")
of_kv(13,"Max Price @ Cash-on-Cash target",
      f"=({MAO['noi']}-{MAO['tcoc']}*{Ucapex})/({MAO['ltv']}*{MAO['k']}+{MAO['tcoc']}*(1-{MAO['ltv']}+{Uclose}))",
      CUR2,"CoC-constrained ceiling",key="pcoc")
of_kv(14,"MAXIMUM ALLOWABLE OFFER (MAO)",f"=MIN({MAO['pdscr']},{MAO['pcoc']})",CUR2,
      "highest price where BOTH clear",big=True,fill=KEY_FILL,key="mao")

TD,TC="$B$5","$B$6"   # target cells (absolute)

# ---- offer comparison -------------------------------------------------------
sec(of, 16, "OFFER COMPARISON  (type any price/structure — verdict updates live)", "A:E")
hdr=17
for cl,t in zip("ABCDE",["Lever / Metric","Seller Asking","Offer 1: Conventional",
                          "Offer 2: Partial Carry","Offer 3: Full Carry"]):
    c=of[f"{cl}{hdr}"]; c.value=t; c.font=SECTION; c.fill=SECTION_FILL
    c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
of.row_dimensions[hdr].height=30

def of_in(row,label,vals,fmt,note=""):
    of[f"A{row}"]=label; of[f"A{row}"].font=LABEL
    for cl,v in zip("BCDE",vals):
        c=of[f"{cl}{row}"]; c.value=v; c.number_format=fmt; c.font=INPUTF; c.fill=INPUT_FILL
        c.border=BORDER; c.alignment=Alignment(horizontal="right")
def of_ca(row,label,tmpl,fmt,bold=False,fill=None):
    of[f"A{row}"]=label; of[f"A{row}"].font=BOLD if bold else LABEL
    for cl in "BCDE":
        c=of[f"{cl}{row}"]; c.value=tmpl.format(c=cl); c.number_format=fmt
        c.font=BOLD if bold else LABEL; c.alignment=Alignment(horizontal="right")
        if fill: c.fill=fill

of_in(18,"Purchase Price",[1500000,1150000,1500000,1500000],CUR2)
of["F18"]="Offer 1 defaults near the MAO above — type your own to test"; of["F18"].font=NOTE
of_in(19,"Bank Loan % (of price)",[0.70,0.70,0.65,0.0],PCT)
of_in(20,"Bank Interest Rate",[0.07,0.07,0.07,0.07],PCT2)
of_in(21,"Bank Amortization (yrs)",[25,25,25,25],CUR)
of_in(22,"Bank Interest-Only? (Y/N)",["N","N","N","N"],'@')
of_in(23,"Seller Carry % (of price)",[0,0,0.25,0.90],PCT)
of_in(24,"Seller Carry Rate",[0,0,0.05,0.06],PCT2)
of_in(25,"Seller Carry Amort (yrs)",[25,25,25,25],CUR)
of_in(26,"Seller Carry Interest-Only? (Y/N)",["N","N","N","Y"],'@')
of_ca(27,"Down Payment %","=1-{c}19-{c}23",PCT)

sec(of, 28, "DEBT STACK", "A:E")
of_ca(29,"Bank Loan $","={c}18*{c}19",CUR2)
of_ca(30,"Bank Annual Debt Service",'=IF({c}22="Y",{c}29*{c}20,PMT({c}20/12,{c}21*12,-{c}29)*12)',CUR2)
of_ca(31,"Seller Carry $","={c}18*{c}23",CUR2)
of_ca(32,"Seller Carry Annual Debt Service",
      '=IF({c}31=0,0,IF({c}26="Y",{c}31*{c}24,PMT({c}24/12,{c}25*12,-{c}31)*12))',CUR2)
of_ca(33,"Total Annual Debt Service","={c}30+{c}32",CUR2,bold=True,fill=TOTAL_FILL)
of_ca(34,"Blended Rate",'=IF(({c}29+{c}31)=0,0,({c}29*{c}20+{c}31*{c}24)/({c}29+{c}31))',PCT2)
of_ca(35,"Down Payment $","={c}18*{c}27",CUR2)
of_ca(36,"Closing + CapEx",f"={{c}}18*{Uclose}+{Ucapex}",CUR2)
of_ca(37,"Total Cash Invested","={c}35+{c}36",CUR2,bold=True,fill=TOTAL_FILL)

sec(of, 38, "KEY METRICS", "A:E")
of_ca(39,"NOI (constant)",f"={Uno}",CUR2)
of_ca(40,"Cap Rate","={c}39/{c}18",PCT)
of_ca(41,"DSCR (total debt)",'=IF({c}33=0,"n/a",{c}39/{c}33)',MULT,bold=True)
of_ca(42,"Net Profit (cash flow)","={c}39-{c}33",CUR2)
of_ca(43,"Cash-on-Cash Return","={c}42/{c}37",PCT,bold=True)
of_ca(44,"DSCR clears target?",f'=IF({{c}}33=0,"YES",IF({{c}}39/{{c}}33>={TD},"YES","NO"))','@')
of_ca(45,"Cash-on-Cash clears target?",f'=IF({{c}}43>={TC},"YES","NO")','@')
of_ca(46,"VERDICT",
      '=IF(AND({c}44="YES",{c}45="YES"),"GO",IF(OR({c}44="YES",{c}45="YES"),"CONDITIONAL","NO-GO"))','@',bold=True)
for cl in "BCDE":
    of[f"{cl}44"].alignment=Alignment(horizontal="center")
    of[f"{cl}45"].alignment=Alignment(horizontal="center")
    of[f"{cl}46"].alignment=Alignment(horizontal="center"); of[f"{cl}46"].font=Font(bold=True,size=12)

GREEN2=PatternFill("solid", fgColor="C6EFCE"); YEL2=PatternFill("solid", fgColor="FFEB9C"); RED2=PatternFill("solid", fgColor="FFC7CE")
GF=Font(color="006100", bold=True); YF=Font(color="9C6500", bold=True); RF=Font(color="9C0006", bold=True)
for rge in ("B44:E45",):
    of.conditional_formatting.add(rge, CellIsRule(operator='equal', formula=['"YES"'], fill=GREEN2, font=GF))
    of.conditional_formatting.add(rge, CellIsRule(operator='equal', formula=['"NO"'],  fill=RED2,  font=RF))
of.conditional_formatting.add("B46:E46", CellIsRule(operator='equal', formula=['"GO"'], fill=GREEN2, font=GF))
of.conditional_formatting.add("B46:E46", CellIsRule(operator='equal', formula=['"CONDITIONAL"'], fill=YEL2, font=YF))
of.conditional_formatting.add("B46:E46", CellIsRule(operator='equal', formula=['"NO-GO"'], fill=RED2, font=RF))

of["A48"]=("Linda's 3 levers to save a CONDITIONAL deal: (1) lower the price toward the MAO, (2) ask the seller "
           "to carry a slice — the blended rate drops so DSCR clears and CoC jumps, (3) add down payment. "
           "Full seller carry (interest-only) often clears both even at the asking price.")
of["A48"].font=NOTE
of["A49"]="Don't pay the seller for value YOU create (rent bumps, occupancy, utility bill-back, new revenue)."
of["A49"].font=NOTE

# =============================================================================
# SHEET — READ ME (insert first)
# =============================================================================
rm = wb.create_sheet("Read Me", 0)
rm.sheet_view.showGridLines=False
rm.column_dimensions["A"].width=104
rm["A1"]="RV PARK UNDERWRITING — HOW TO USE"
rm["A1"].font=TITLE; rm["A1"].fill=TITLE_FILL
rm["A1"].alignment=Alignment(horizontal="left", vertical="center", indent=1)
rm.row_dimensions[1].height=26
lines=[
 "",
 "WHAT THIS IS",
 "An RV-park deal-screening model built on the LindaAI underwriting framework. Drop in a",
 "deal's numbers and it tells you in minutes whether it makes money — using Linda's 5 key metrics, its 5",
 "quick hacks, NOI normalization, and its go / conditional / no-go rule. Works for RV parks and adapts",
 "directly to mobile home parks (same lot-rent math). Part of the LindaAI underwriting framework.",
 "",
 "LINDA'S 5 KEY METRICS (the heartbeat of the deal)",
 "  1. NOI — effective gross income (EGI) minus operating expenses. Every other metric flows from it.",
 "  2. Cap Rate = NOI / price. Higher = paying less. But a very high cap (14%+) usually means hidden risk.",
 "  3. DSCR = NOI / annual debt. Linda's floor is 1.35x — that's YOUR target, not just the lender's.",
 "  4. Net Profit = what's left after opex AND debt service.",
 "  5. Cash-on-Cash = net profit / cash invested. Linda's target is 10%+.",
 "",
 "LINDA'S 5 QUICK HACKS (screen any listing in ~2 minutes) — see the Deal Scorecard tab",
 "  1. 1% Rule — monthly gross >= 1% of price = likely good (cross the last two zeros off the price).",
 "  2. 10x Rule (GRM) — price / annual gross; want UNDER 10. Over ~15 is expensive.",
 "  3. Quick DSCR — NOI / (price x interest rate): >1.5 good, 1.25-1.5 closer look, <1.25 lending risk.",
 "  4. Expense Ratio — OpEx / gross: UNDER 30% = numbers are fiction; 35-65% normal; over 70% a problem.",
 "  5. Negative Leverage — if cap rate < your interest rate you lose money day one. Want cap ABOVE rate.",
 "",
 "NORMALIZE THE NOI (Linda: never trust the seller's NOI) — see the Normalization tab",
 "  The OM is the seller's best case; the Pro Forma is fantasy; the T12 is your start but needs adjusting.",
 "  Linda's 4 standard adjustments: management -> 10% of EGI, repairs -> 5% (or prior year), capex -> +3%,",
 "  property taxes -> +20% (they reset on sale). A small NOI cut at an 8-10% cap = a huge price difference.",
 "",
 "THE BAD-DAY TEST (break the deal on purpose) — use the Sensitivity tab",
 "  Cut occupancy 10-15%, add 1-2% to the interest rate, raise expenses 10-15%. Does DSCR still clear?",
 "",
 "THE DECISION (after normalizing)",
 "  Both DSCR (>=1.35) AND Cash-on-Cash (>=10%) clear -> GO. One fails -> CONDITIONAL (restructure).",
 "  Both fail -> likely NO-GO. To save a conditional deal, use Linda's 3 levers: lower the price to your",
 "  MAO, ask the seller to carry a slice (blended rate drops), or add down payment. Don't pay the seller",
 "  for value YOU will create (rent bumps to market, occupancy, utility bill-back, added revenue).",
 "",
 "THE TABS (left to right)",
 "  • Deal Summary  — one-page headline view; print/screenshot for lenders & partners.",
 "  • Deal Scorecard — Linda's metrics + 5 hacks rated GOOD/OK/BAD with its go / conditional / no-go verdict.",
 "  • Underwriting  — the engine. Fill the YELLOW cells; everything else is a formula.",
 "  • Normalization — rebuild the seller's NOI with Linda's 4 adjustments; shows the overpayment risk.",
 "  • Offer Structures — Linda's 4-column engine: Seller / Conventional / Partial carry / Full carry, two-lien",
 "      debt stack, blended rate, and the auto-solved Maximum Allowable Offer (MAO).",
 "  • Income Detail — optional. Build income from your site mix (long-term + seasonal nightly).",
 "  • Actuals vs Pro Forma — seller's CURRENT numbers vs your stabilized plan, side by side.",
 "  • Pro Forma & Returns — 10-year projection, sale reversion, IRR & equity multiple.",
 "  • Sensitivity   — the Bad-Day Test: CoC & DSCR vs price, interest rate, rent & vacancy.",
 "  • Loan Sizing   — max purchase price the NOI supports at your target DSCR / debt yield.",
 "",
 "HOW TO USE IT",
 "1. Normalization tab: enter the seller's reported expenses; read off your normalized NOI.",
 "2. Underwriting tab: fill the yellow cells using those NORMALIZED numbers, not the seller's.",
 "3. Deal Scorecard: read the GOOD/OK/BAD ratings and the GO / CONDITIONAL / NO-GO verdict.",
 "4. Sensitivity: run the bad-day stress test. Deal Summary: the one-page headline.",
 "",
 "RV-PARK / MHP WATCH-OUTS",
 "  • Nondisclosure states: sellers needn't volunteer problems (failing septic, AC, zoning). Dig.",
 "  • Transient/nightly income is seasonal and far less stable than annual/monthly tenants.",
 "  • Utility metering — are sites individually metered, or is the owner eating utilities? (ratio bill-back).",
 "  • Infrastructure age — septic/sewer, electric pedestals (30/50 amp), water lines = big CapEx.",
 "  • Property taxes usually RESET on sale — don't trust the seller's current tax line.",
 "",
 "This is a screening tool, not investment advice. Verify every number and consult professionals.",
]
r=2
for ln in lines:
    rm[f"A{r}"]=ln
    if ln.isupper() and ln.strip() and not ln.startswith(" "):
        rm[f"A{r}"].font=Font(bold=True, size=12, color="1F4E78")
    else:
        rm[f"A{r}"].font=Font(size=11)
    r+=1

# ---- tab order --------------------------------------------------------------
order=["Read Me","Deal Summary","Deal Scorecard","Underwriting","Normalization","Offer Structures",
       "Income Detail","Actuals vs Pro Forma","Pro Forma & Returns","Sensitivity","Loan Sizing"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = 0

wb.save("RV_Park_Underwriting.xlsx")
print("Saved RV_Park_Underwriting.xlsx")
print("Tabs:", [s.title for s in wb._sheets])
